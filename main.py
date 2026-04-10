import io
import json
import os
import secrets
from typing import Annotated, List

import openpyxl
from fastapi import Depends, FastAPI, Form, HTTPException, Request, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.templating import Jinja2Templates

from database import (
    init_db,
    migrar_fechas_a_chile,
    obtener_stats,
    insertar_punto_compra,
    obtener_todas_punto_compra,
    eliminar_punto_compra,
    insertar_visita,
    insertar_entrevistas_visita,
    obtener_todas_visitas,
    obtener_entrevistas_visita,
    eliminar_visita,
    limpiar_visitas,
    limpiar_punto_compra,
)

app = FastAPI(title="En los zapatos del cliente")
templates = Jinja2Templates(directory="templates")
security = HTTPBasic()

# Version: 2.1.0 - Fix error 500 con fallback a campos basicos

# Credenciales del dashboard (configurar en Render como env vars)
DASHBOARD_USER = os.getenv("DASHBOARD_USER", "admin")
DASHBOARD_PASS = os.getenv("DASHBOARD_PASS", "walmart2025")


@app.on_event("startup")
def on_startup() -> None:
    init_db()
    resultado = migrar_fechas_a_chile()
    print(f"[TZ Migration] {resultado}")


# --- Autenticacion basica para el dashboard ---
def verificar_credenciales(
    credentials: Annotated[HTTPBasicCredentials, Depends(security)],
) -> str:
    user_ok = secrets.compare_digest(credentials.username, DASHBOARD_USER)
    pass_ok = secrets.compare_digest(credentials.password, DASHBOARD_PASS)
    if not (user_ok and pass_ok):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Credenciales incorrectas",
            headers={"WWW-Authenticate": "Basic"},
        )
    return credentials.username


# --- Redireccion raiz ---
@app.get("/", response_class=HTMLResponse)
def raiz(request: Request):
    return RedirectResponse(url="/visitas", status_code=302)


# --- Formulario Punto de Compra ---
@app.get("/punto-compra", response_class=HTMLResponse)
def mostrar_punto_compra(request: Request):
    """Formulario de evaluación Punto de Compra."""
    return templates.TemplateResponse("punto_compra.html", {"request": request})


@app.post("/submit-punto-compra")
def recibir_punto_compra(
    nombre: Annotated[str, Form()] = "",
    tienda: Annotated[str, Form()] = "",
    # Step 1: Entrada al local
    s1_i1: Annotated[str, Form()] = "",
    s1_obs: Annotated[str, Form()] = "",
    # Step 2: Vitrinear
    s2_i1: Annotated[str, Form()] = "",
    s2_i2: Annotated[str, Form()] = "",
    s2_i3: Annotated[str, Form()] = "",
    s2_obs: Annotated[str, Form()] = "",
    # Step 3: Selección
    s3_i1: Annotated[str, Form()] = "",
    s3_i4: Annotated[str, Form()] = "",
    s3_i3: Annotated[str, Form()] = "",
    s3_obs: Annotated[str, Form()] = "",
    # Step 4: Pago
    s4_i1: Annotated[str, Form()] = "",
    s4_i2: Annotated[str, Form()] = "",
    s4_obs: Annotated[str, Form()] = "",
    # Step 5: Espera
    s5_obs: Annotated[str, Form()] = "",
):
    """Recibe la evaluación de punto de compra."""
    insertar_punto_compra({
        "nombre": nombre,
        "tienda": tienda,
        "s1_encontrar_punto": s1_i1,
        "s1_observaciones": s1_obs,
        "s2_vitrinear": s2_i1,
        "s2_comparar": s2_i2,
        "s2_autonomo": s2_i3,
        "s2_observaciones": s2_obs,
        "s3_agregar_carro": s3_i1,
        "s3_crear_usuario": s3_i4,
        "s3_despacho_retiro": s3_i3,
        "s3_observaciones": s3_obs,
        "s4_medio_pago": s4_i1,
        "s4_proceso_pago": s4_i2,
        "s4_observaciones": s4_obs,
        "s5_observaciones": s5_obs,
    })
    return RedirectResponse(url="/gracias-punto-compra", status_code=303)


@app.get("/gracias-punto-compra", response_class=HTMLResponse)
def gracias_punto_compra(request: Request):
    return templates.TemplateResponse("gracias_punto_compra.html", {"request": request})


@app.get("/gracias", response_class=HTMLResponse)
def gracias_visita(request: Request):
    return templates.TemplateResponse("gracias.html", {"request": request})


# --- Formulario Visitas con Sentido ---
@app.get("/visitas", response_class=HTMLResponse)
def mostrar_visitas(request: Request):
    """Formulario Visitas con Sentido."""
    return templates.TemplateResponse("visitas.html", {"request": request})


@app.post("/submit-visita")
async def recibir_visita(
    request: Request,
):
    """Recibe el formulario Visitas con Sentido v2."""
    print("="*60)
    print("[INICIO] Recibiendo formulario visitas")
    
    # Obtener TODOS los datos del formulario
    form_data = await request.form()
    
    print(f"[DEBUG] TODOS LOS CAMPOS RECIBIDOS ({len(form_data)} total):")
    for key, value in form_data.items():
        if len(str(value)) > 100:
            print(f"  {key} = {str(value)[:100]}... (truncado)")
        else:
            print(f"  {key} = {value}")
    print("="*60)
    
    # Extraer campos individuales
    geo_lat = form_data.get("geo_lat", "")
    geo_lng = form_data.get("geo_lng", "")
    usuario = form_data.get("usuario", "")
    q1_otro_text = form_data.get("q1_otro_text", "")
    local = form_data.get("local", "")
    q2_otro_text = form_data.get("q2_otro_text", "")
    tiempo_fila = form_data.get("tiempo_fila", "00:00")
    
    # Paso 1: Atención
    q3 = form_data.get("q3", "")
    q3_otro_text = form_data.get("q3_otro_text", "")
    q4 = form_data.get("q4", "")
    q4_otro_text = form_data.get("q4_otro_text", "")
    q5 = form_data.get("q5", "")
    q5_otro_text = form_data.get("q5_otro_text", "")
    q6 = form_data.get("q6", "")
    q6_otro_text = form_data.get("q6_otro_text", "")
    q7 = form_data.get("q7", "")
    q7_otro_text = form_data.get("q7_otro_text", "")
    q8_csat = form_data.get("q8_csat", "")
    q9_carteleria = form_data.get("q9_carteleria", "")
    q9_carteleria_otro_text = form_data.get("q9_carteleria_otro_text", "")
    
    # Paso 2: Zona de pago
    q8 = form_data.get("q8", "")
    q9 = form_data.get("q9", "")
    q10 = form_data.get("q10", "")
    q11 = form_data.get("q11", "")
    q12 = form_data.get("q12", "")
    q13 = form_data.get("q13", "")
    
    # Paso 3: Entrevistas
    entrevista_1_motivo = form_data.get("entrevista_1_motivo", "")
    entrevista_1_positivos = form_data.get("entrevista_1_positivos", "")
    entrevista_1_mejoras = form_data.get("entrevista_1_mejoras", "")
    
    # Comentarios finales
    q17 = form_data.get("q17", "")
    
    print(f"[DATA] usuario={usuario}, local={local}")
    print(f"[DATA] q3={q3}, q7={q7}, q9_carteleria={q9_carteleria}")
    print(f"[DATA ENTREVISTAS]:")
    print(f"  entrevista_1_motivo = '{entrevista_1_motivo}' (len={len(entrevista_1_motivo)})")
    print(f"  entrevista_1_positivos = '{entrevista_1_positivos}' (len={len(entrevista_1_positivos)})")
    print(f"  entrevista_1_mejoras = '{entrevista_1_mejoras}' (len={len(entrevista_1_mejoras)})")
    print("="*60)
    
    try:
        # Si seleccionó "otro" en usuario o local, usar el texto libre
        usuario_final = q1_otro_text.strip() if usuario == "otro" and q1_otro_text.strip() else usuario
        local_final = q2_otro_text.strip() if local == "otro" and q2_otro_text.strip() else local
        
        visita_id = insertar_visita({
            "geo_lat": geo_lat,
            "geo_lng": geo_lng,
            "usuario": usuario_final,
            "local": local_final,
            "tiempo_fila": tiempo_fila,
            "q3": q3, "q3_otro_text": q3_otro_text,
            "q4": q4, "q4_otro_text": q4_otro_text,
            "q5": q5, "q5_otro_text": q5_otro_text,
            "q6": q6, "q6_otro_text": q6_otro_text,
            "q7": q7, "q7_otro_text": q7_otro_text,
            "q8_csat": q8_csat,
            "q9_carteleria": q9_carteleria,
            "q9_carteleria_otro_text": q9_carteleria_otro_text,
            "q8": q8,
            "q9": q9, "q10": q10, "q11": q11, "q12": q12, "q13": q13,
            "q17": q17,
        })
        
        # Insertar entrevistas si tienen contenido
        entrevistas = []
        if entrevista_1_motivo.strip() or entrevista_1_positivos.strip() or entrevista_1_mejoras.strip():
            entrevistas.append({
                "numero_cliente": 1,
                "motivo_visita": entrevista_1_motivo.strip(),
                "aspectos_positivos": entrevista_1_positivos.strip(),
                "oportunidades_mejora": entrevista_1_mejoras.strip(),
            })
        
        if entrevistas:
            insertar_entrevistas_visita(visita_id, entrevistas)
            print(f"[OK] Insertadas {len(entrevistas)} entrevista(s) para visita_id={visita_id}")
        
        return RedirectResponse(url="/gracias", status_code=303)
    except Exception as e:
        print(f"ERROR al insertar visita: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error al guardar: {str(e)}")



@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(
    request: Request,
    _: Annotated[str, Depends(verificar_credenciales)],
    tab: str = "visitas",
):
    # Obtener datos de punto de compra
    punto_compra_rows = obtener_todas_punto_compra()

    # Obtener datos de Visitas con Sentido
    visita_rows = obtener_todas_visitas()
    visitas_con_entrevistas = []
    for row in visita_rows:
        entrevistas = obtener_entrevistas_visita(row["id"])
        visitas_con_entrevistas.append({**row, "entrevistas": entrevistas})

    stats = obtener_stats()
    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "rows": visitas_con_entrevistas,
            "punto_compra_rows": punto_compra_rows,
            "visita_rows": visitas_con_entrevistas,
            "stats": stats,
            "tab": tab,
        },
    )


@app.post("/dashboard/delete-punto-compra/{row_id}")
def eliminar_pc(
    row_id: int,
    _: Annotated[str, Depends(verificar_credenciales)],
):
    """Elimina una evaluación de punto de compra."""
    eliminar_punto_compra(row_id)
    return RedirectResponse(url="/dashboard?tab=punto-compra", status_code=303)


@app.post("/dashboard/delete-visita/{row_id}")
def eliminar_visita_endpoint(
    row_id: int,
    _: Annotated[str, Depends(verificar_credenciales)],
):
    """Elimina una visita con sentido."""
    eliminar_visita(row_id)
    return RedirectResponse(url="/dashboard?tab=visitas", status_code=303)


@app.post("/dashboard/limpiar-visitas")
def limpiar_visitas_endpoint(
    _: Annotated[str, Depends(verificar_credenciales)],
):
    """Limpia TODOS los registros de Visitas con Sentido."""
    limpiar_visitas()
    return RedirectResponse(url="/dashboard?tab=visitas", status_code=303)


@app.post("/dashboard/limpiar-punto-compra")
def limpiar_punto_compra_endpoint(
    _: Annotated[str, Depends(verificar_credenciales)],
):
    """Limpia TODOS los registros de Punto de Compra."""
    limpiar_punto_compra()
    return RedirectResponse(url="/dashboard?tab=punto-compra", status_code=303)


@app.get("/admin/reset-db")
def reset_database(
    username: Annotated[str, Depends(verificar_credenciales)]
):
    """Endpoint para recrear la tabla visitas desde cero."""
    try:
        from database import get_conn
        
        with get_conn() as conn:
            # 1. Eliminar tabla vieja
            print("[RESET] Eliminando tabla visitas...")
            conn.execute("DROP TABLE IF EXISTS visitas CASCADE")
            conn.commit()
            
            # 2. Crear tabla nueva con TODAS las columnas
            print("[RESET] Creando tabla visitas con todas las columnas...")
            sql = """
            CREATE TABLE visitas (
                id SERIAL PRIMARY KEY,
                fecha TEXT NOT NULL,
                geo_lat TEXT,
                geo_lng TEXT,
                usuario TEXT,
                local TEXT,
                tiempo_fila TEXT,
                q3 TEXT,
                q3_otro_text TEXT,
                q4 TEXT,
                q4_otro_text TEXT,
                q5 TEXT,
                q5_otro_text TEXT,
                q6 TEXT,
                q6_otro_text TEXT,
                q7 TEXT,
                q7_otro_text TEXT,
                q8_csat TEXT,
                q9_carteleria TEXT,
                q9_carteleria_otro_text TEXT,
                q8 TEXT,
                q9 TEXT,
                q10 TEXT,
                q11 TEXT,
                q12 TEXT,
                q13 TEXT,
                q17 TEXT
            )
            """
            conn.execute(sql)
            conn.commit()
            
            # 3. Verificar columnas creadas
            print("[RESET] Verificando columnas creadas...")
            result = conn.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = 'visitas'
                ORDER BY ordinal_position
            """).fetchall()
            
            columnas = [r['column_name'] for r in result]
            print(f"[RESET] Columnas creadas: {columnas}")
            
            return {
                "success": True,
                "message": "Tabla visitas recreada exitosamente",
                "columnas_creadas": columnas,
                "total_columnas": len(columnas)
            }
            
    except Exception as e:
        print(f"[ERROR] Error al recrear tabla: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/admin/debug-bd")
def debug_bd(username: Annotated[str, Depends(verificar_credenciales)]):
    """Endpoint temporal para ver qué se guardó en BD (DEBUG)."""
    from database import get_conn
    
    html = "<html><head><meta charset='utf-8'><title>Debug BD</title><style>"
    html += "body { font-family: monospace; padding: 20px; background: #f5f5f5; }"
    html += "table { border-collapse: collapse; background: white; width: 100%; }"
    html += "th, td { border: 1px solid #ddd; padding: 8px; text-align: left; font-size: 12px; }"
    html += "th { background: #0053e2; color: white; position: sticky; top: 0; }"
    html += ".vacio { color: red; font-weight: bold; } .lleno { color: green; }"
    html += "</style></head><body>"
    html += "<h1>🔍 Debug BD - Última Visita</h1>"
    
    with get_conn() as conn:
        # Columnas
        columnas = conn.execute("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'visitas'
            ORDER BY ordinal_position
        """).fetchall()
        
        html += f"<h2>✅ Columnas en tabla visitas ({len(columnas)} total)</h2>"
        html += "<ol style='columns: 3;'>" 
        for col in columnas:
            html += f"<li>{col['column_name']}</li>"
        html += "</ol>"
        
        # Última visita
        visita = conn.execute("""
            SELECT * FROM visitas ORDER BY id DESC LIMIT 1
        """).fetchone()
        
        if visita:
            html += f"<h2>📊 Última Visita (ID {visita['id']})</h2>"
            html += "<table>"
            html += "<tr><th>Campo</th><th>Valor</th><th>Estado</th></tr>"
            
            for col in columnas:
                campo = col['column_name']
                valor = visita.get(campo, '')
                if valor:
                    estado = '<span class="lleno">✅ Lleno</span>'
                    valor_mostrar = str(valor)[:100]  # Truncar si es muy largo
                else:
                    estado = '<span class="vacio">❌ Vacío</span>'
                    valor_mostrar = '(vacío)'
                html += f"<tr><td><strong>{campo}</strong></td><td>{valor_mostrar}</td><td>{estado}</td></tr>"
            
            html += "</table>"
        else:
            html += "<p style='color:red; font-size: 20px;'>⚠️ NO HAY VISITAS EN LA BD</p>"
            html += "<p>Resetea la BD y llena el formulario primero.</p>"
    
    html += "<br><br><a href='/dashboard' style='color: #0053e2; font-size: 18px;'>← Volver al Dashboard</a>"
    html += "</body></html>"
    
    return HTMLResponse(content=html)


@app.get("/exportar-excel")
def exportar_excel(
    _: Annotated[str, Depends(verificar_credenciales)],
):
    """Exporta visitas y punto de compra a Excel."""
    wb = openpyxl.Workbook()

    # Hoja 1: Visitas con Sentido
    ws1 = wb.active
    ws1.title = "Visitas con Sentido"
    headers1 = [
        "ID", "Fecha", "Usuario", "Local", "Geo Lat", "Geo Lng", "Tiempo Fila",
        "Q3: Guardia saludó", "Q3 Otro",
        "Q4: Guardia preguntó", "Q4 Otro",
        "Q5: Pasillos saludo", "Q5 Otro",
        "Q6: Pasillos preguntaron", "Q6 Otro",
        "Q7: Encontró colaborador", "Q7 Otro",
        "Q8: CSAT Resolutividad (1-7)",
        "Q9: Cartelería TyC", "Q9 Otro",
        "Q8: Tipo Cajero",
        "Q9: Cajero saludó", "Q10: PMC", "Q11: Líder BCI", 
        "Q12: Boleta Mail", "Q13: Despedida",
        "Q17: Comentarios Finales",
    ]
    ws1.append(headers1)
    for r in obtener_todas_visitas():
        ws1.append([
            r["id"], r["fecha"], r["usuario"], r["local"],
            r.get("geo_lat", ""), r.get("geo_lng", ""), r.get("tiempo_fila", "--:--"),
            r.get("q3", ""), r.get("q3_otro_text", ""),
            r.get("q4", ""), r.get("q4_otro_text", ""),
            r.get("q5", ""), r.get("q5_otro_text", ""),
            r.get("q6", ""), r.get("q6_otro_text", ""),
            r.get("q7", ""), r.get("q7_otro_text", ""),
            r.get("q8_csat", ""),
            r.get("q9_carteleria", ""), r.get("q9_carteleria_otro_text", ""),
            r.get("q8", ""),
            r.get("q9", ""), r.get("q10", ""), r.get("q11", ""),
            r.get("q12", ""), r.get("q13", ""),
            r.get("q17", ""),
        ])

    # Hoja 2: Entrevistas de Visitas
    ws2 = wb.create_sheet(title="Entrevistas Visitas")
    headers2 = [
        "Visita ID", "Usuario", "Local", "# Cliente",
        "Motivo Visita", "Aspectos Positivos", "Oportunidades Mejora",
    ]
    ws2.append(headers2)
    for r in obtener_todas_visitas():
        for e in obtener_entrevistas_visita(r["id"]):
            ws2.append([
                r["id"], r["usuario"], r["local"],
                e["numero_cliente"],
                e.get("motivo_visita", ""),
                e.get("aspectos_positivos", ""),
                e.get("oportunidades_mejora", ""),
            ])

    # Hoja 3: Punto de Compra
    ws3 = wb.create_sheet(title="Punto de Compra")
    headers3 = [
        "ID", "Fecha", "Nombre", "Tienda",
        "S1 Encontrar Punto", "S1 Obs",
        "S2 Vitrinear", "S2 Comparar", "S2 Autónomo", "S2 Obs",
        "S3 Agregar Carro", "S3 Crear Usuario", "S3 Despacho/Retiro", "S3 Obs",
        "S4 Medio Pago", "S4 Proceso Pago", "S4 Obs",
        "S5 Obs",
    ]
    ws3.append(headers3)
    for r in obtener_todas_punto_compra():
        ws3.append([
            r["id"], r["fecha"], r["nombre"], r["tienda"],
            r.get("s1_encontrar_punto", ""), r.get("s1_observaciones", ""),
            r.get("s2_vitrinear", ""), r.get("s2_comparar", ""), r.get("s2_autonomo", ""), r.get("s2_observaciones", ""),
            r.get("s3_agregar_carro", ""), r.get("s3_crear_usuario", ""), r.get("s3_despacho_retiro", ""), r.get("s3_observaciones", ""),
            r.get("s4_medio_pago", ""), r.get("s4_proceso_pago", ""), r.get("s4_observaciones", ""),
            r.get("s5_observaciones", ""),
        ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=visitas_export.xlsx"},
    )
